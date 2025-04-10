import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont   # Use InlineFont instead of Font for rich text
from io import BytesIO
from datetime import datetime

# --- Helper functions for date formatting ---

def ordinal(n):
    """Return the ordinal string of a number (e.g., 9 -> '9th')."""
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return str(n) + suffix

def format_date(val):
    """If a date is in the format '09-04-2025  05:00:00', format it as '9th April 2025'."""
    if pd.isna(val):
        return ""
    if isinstance(val, datetime):
        dt = val
    else:
        try:
            # Strip extra spaces and parse the input string.
            dt = datetime.strptime(str(val).strip(), "%d-%m-%Y %H:%M:%S")
        except Exception:
            # If parsing fails, return the original string.
            return str(val)
    return f"{ordinal(dt.day)} {dt.strftime('%B')} {dt.year}"

# --- Main function to generate the formatted Excel file ---
def generate_formatted_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Output Final"
    
    # Process each record from the input file (one row per record)
    for idx, row in df.iterrows():
        output_row = idx + 1  # Excel rows are 1-indexed

        # --- COLUMN 1: Record Details ---
        # Line 1: Date line (PlannedStart - PlannedEnd) in bold.
        planned_start = format_date(row['PlannedStart']) if pd.notna(row['PlannedStart']) else ""
        planned_end   = format_date(row['PlannedEnd'])   if pd.notna(row['PlannedEnd'])   else ""
        date_line = f"{planned_start} - {planned_end}".strip()
        
        # Line 2: Title
        title_line = str(row['Title']) if pd.notna(row['Title']) else ""
        
        # Line 3: Summary line (Location, OnLine/Outage, CI (count), BC (count), NONBC (count))
        location = str(row['Location']) if pd.notna(row['Location']) else ""
        online_outage = str(row['OnLine/Outage']) if pd.notna(row['OnLine/Outage']) else ""
        ci_count = len(str(row['CI']).split(",")) if pd.notna(row['CI']) else 0
        bc_count = len(str(row['BC']).split(",")) if pd.notna(row['BC']) else 0
        nonbc_count = len(str(row['NONBC']).split(",")) if pd.notna(row['NONBC']) else 0
        summary_line = f"{location}, {online_outage}, CI ({ci_count} CIs), BC ({bc_count} BC), NONBC ({nonbc_count} NONBC)".strip()
        
        # Line 4: BusinessGroups
        business_groups_line = str(row['BusinessGroups']) if pd.notna(row['BusinessGroups']) else ""
        
        # Build rich text for Column 1.
        # Insert a blank line (two newline characters) after each line.
        col1_rich = CellRichText(
            TextBlock(date_line, InlineFont(bold=True)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock(title_line, InlineFont(bold=False)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock(summary_line, InlineFont(bold=False)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock(business_groups_line, InlineFont(bold=False))
        )
        
        # --- COLUMN 2: Change & Risk ---
        # Use the F4F column if available; otherwise use ChangeId with "/F4F" appended.
        if 'F4F' in row.index:
            f4f_val = str(row['F4F']) if pd.notna(row['F4F']) else ""
        else:
            change_id = str(row['ChangeId']) if pd.notna(row['ChangeId']) else ""
            f4f_val = f"{change_id}/F4F" if change_id else ""
            
        # Process the RiskLevel value (remove a leading SHELL_ if present, then capitalize)
        risk = str(row['RiskLevel']) if pd.notna(row['RiskLevel']) else ""
        if risk.upper().startswith("SHELL_"):
            risk = risk[6:]
        risk = risk.capitalize()
        
        col2_rich = CellRichText(
            TextBlock(f4f_val, InlineFont(bold=False)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock(risk, InlineFont(bold=False))
        )
        
        # --- COLUMN 3: Trading Assets & BC Apps ---
        # Parse the BC column items (assumed comma-separated) that include "(RelationType = Direct)"
        trading_apps = []
        other_apps = []
        if pd.notna(row['BC']):
            for item in str(row['BC']).split(","):
                item = item.strip()
                if "(RelationType = Direct)" in item:
                    app_name = item.replace("(RelationType = Direct)", "").strip()
                    if app_name.upper().startswith("ST"):
                        trading_apps.append(app_name)
                    else:
                        other_apps.append(app_name)
        
        trading_scope = "Yes" if trading_apps else "No"
        trading_bc_apps_content = ", ".join(trading_apps) if trading_apps else "None"
        # For Other BC Apps:
        if not trading_apps:
            other_bc_apps_content = "No"
        else:
            other_bc_apps_content = ", ".join(other_apps) if other_apps else "None"
        
        col3_rich = CellRichText(
            TextBlock("Trading assets in scope: ", InlineFont(bold=True)),
            TextBlock(trading_scope, InlineFont(bold=False)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock("Trading BC Apps: ", InlineFont(bold=True)),
            TextBlock(trading_bc_apps_content, InlineFont(bold=False)),
            TextBlock("\n\n", InlineFont(bold=False)),
            TextBlock("Other BC Apps: ", InlineFont(bold=True)),
            TextBlock(other_bc_apps_content, InlineFont(bold=False))
        )
        
        # Write the rich text objects into their respective cells.
        ws.cell(row=output_row, column=1).value = col1_rich
        ws.cell(row=output_row, column=2).value = col2_rich
        ws.cell(row=output_row, column=3).value = col3_rich
        
        # Set cells to wrap text so multiple lines are visible.
        ws.cell(row=output_row, column=1).alignment = ws.cell(row=output_row, column=1).alignment.copy(wrapText=True)
        ws.cell(row=output_row, column=2).alignment = ws.cell(row=output_row, column=2).alignment.copy(wrapText=True)
        ws.cell(row=output_row, column=3).alignment = ws.cell(row=output_row, column=3).alignment.copy(wrapText=True)
    
    # Save workbook to a BytesIO stream.
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream

# --- Streamlit App UI ---
st.title("Change Formatter App")

uploaded_file = st.file_uploader("Upload your Changes Excel file", type=["xlsx", "xls"])

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        st.subheader("Preview of Input Data")
        st.dataframe(df.head())
        formatted_excel = generate_formatted_excel(df)
        st.download_button(
            label="📥 Download Formatted Output",
            data=formatted_excel,
            file_name="output_final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Error processing file: {e}")
